from django.views.generic.list import ListView
from django.shortcuts import render, redirect, resolve_url
from django.core import serializers
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.conf import settings
from django.contrib import messages
from urllib.parse import urlparse
import math
import re
import csv
import io
import unicodedata

from dal import autocomplete

from MelissaTrasporti.geo import trova_comune
from MelissaTrasporti.forms import ExcelUploadForm
from MelissaTrasporti.models import Fornitore, Commessa, Comune, OffertaCommessa

import folium
import numpy as np
import geopy as geo

from geopy.geocoders import Nominatim
geolocator = Nominatim(user_agent="melissa_trasporti_inc")

CODICE_GENERICO_KEYS = ('codice',)
CODICE_OFFERTA_KEYS = (
    'cod_offerta',
    'codice_offerta',
    'offerta',
    'cod_osb',
    'codice_osb',
    'osb',
    'cod_ocb',
    'codice_ocb',
    'ocb',
)
CODICE_COMMESSA_KEYS = (
    'cod_commessa',
    'codice_commessa',
    'commessa',
    'rif_cc_co',
    'cc_co',
)
PRODUTTORE_KEYS = ('produttore', 'cliente', 'ragione_sociale')
QUANTITA_KEYS = ('quantita', 'qta', 'q_ta')
TIPOLOGIA_KEYS = ('tipologia', 'tipo')
LOCALITA_KEYS = ('localita', 'comune', 'paese', 'luogo_ritiro')
PROVINCIA_KEYS = ('provincia', 'prov')
GARANZIA_KEYS = ('garanzia_fin', 'garanzia', 'garanzia_finanziaria')
IS_COMMESSA_KEYS = ('is_commessa', 'check_commessa')
IS_DONE_KEYS = ('is_done', 'done', 'evasa', 'commessa_evasa')
PROCESSO_COMPLETATO_KEYS = ('processo_completato', 'processo_completo')
HEADER_FIELD_ALIASES = (
    CODICE_GENERICO_KEYS + CODICE_OFFERTA_KEYS + CODICE_COMMESSA_KEYS,
    PRODUTTORE_KEYS,
    QUANTITA_KEYS,
    TIPOLOGIA_KEYS,
    LOCALITA_KEYS,
    PROVINCIA_KEYS,
    GARANZIA_KEYS,
)


def _normalizza_colonna(nome):
    testo = str(nome or '').strip().lower()
    testo = unicodedata.normalize('NFKD', testo)
    testo = ''.join(char for char in testo if not unicodedata.combining(char))
    testo = testo.replace("'", '')
    return re.sub(r'[^a-z0-9]+', '_', testo).strip('_')


def _valore(record, *nomi):
    for nome in nomi:
        valore = record.get(nome)
        if valore is not None and not (isinstance(valore, float) and math.isnan(valore)):
            testo = str(valore).strip()
            if testo and testo.lower() != 'nan':
                return valore
    return None


def _bool_import(record, *nomi, default=False):
    if nomi and isinstance(nomi[-1], bool):
        default = nomi[-1]
        nomi = nomi[:-1]
    valore = _valore(record, *nomi)
    if valore is None:
        return default
    if isinstance(valore, bool):
        return valore
    return str(valore).strip().lower() in {'1', 'true', 'vero', 'si', 's', 'yes', 'y', 'x'}


def _float_import(record, *nomi):
    valore = _valore(record, *nomi)
    if valore is None:
        return None
    try:
        return float(str(valore).replace(',', '.'))
    except ValueError:
        return None


def _int_import(record, *nomi):
    valore = _valore(record, *nomi)
    if valore is None:
        return None
    if isinstance(valore, (int, float)) and not (isinstance(valore, float) and math.isnan(valore)):
        return int(float(valore))

    testo = str(valore).strip()
    if not testo or testo.lower() == 'nan':
        return None
    testo = re.sub(r'[^\d,.\-]+', '', testo)
    if not testo:
        return None

    if ',' in testo and '.' in testo and testo.rfind(',') > testo.rfind('.'):
        testo = testo.replace('.', '').replace(',', '.')
    elif ',' in testo:
        testo = testo.replace(',', '.')
    elif testo.count('.') > 1:
        testo = testo.replace('.', '')

    try:
        return int(float(testo))
    except ValueError:
        return None


def _normalizza_garanzia(record):
    valore = _valore(record, *GARANZIA_KEYS)
    testo = str(valore or '-').strip().upper()
    if not testo or testo == 'NAN':
        return '-'
    if 'ANTE' in testo:
        return 'ANTE'
    if 'REM' in testo or 'ERION' in testo:
        return 'REM/ERION'
    return '-'


def _inferisci_tipo_import(*parti):
    testo = _normalizza_colonna(' '.join(str(parte or '') for parte in parti))
    ha_offerte = 'offert' in testo
    ha_commesse = 'commess' in testo
    if ha_offerte and not ha_commesse:
        return 'offerta'
    if ha_commesse and not ha_offerte:
        return 'commessa'
    if ha_offerte:
        return 'offerta'
    return None


def _header_score(headers):
    normalizzati = [_normalizza_colonna(header) for header in headers]
    normalizzati = [header for header in normalizzati if header]
    if not normalizzati:
        return 0

    trovati = sum(
        1
        for aliases in HEADER_FIELD_ALIASES
        if any(header in aliases for header in normalizzati)
    )
    if trovati < 4:
        return 0
    return trovati * 10 + min(len(normalizzati), 30)


def _normalizza_headers(headers):
    normalizzati = []
    conteggi = {}
    for index, header in enumerate(headers, start=1):
        nome = _normalizza_colonna(header) or f'col_{index}'
        conteggi[nome] = conteggi.get(nome, 0) + 1
        if conteggi[nome] > 1:
            nome = f'{nome}_{conteggi[nome]}'
        normalizzati.append(nome)
    return normalizzati


def _trova_riga_header(sheet):
    best_score = 0
    best_row_number = None
    best_headers = None

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row_number > 40:
            break
        score = _header_score(row)
        if score > best_score:
            best_score = score
            best_row_number = row_number
            best_headers = row

    if not best_headers:
        return None, None
    return best_row_number, _normalizza_headers(best_headers)


def _riga_vuota(row):
    return all(value is None or not str(value).strip() for value in row)


def _record_utilizzabile(record):
    return any(
        _valore(record, *aliases) is not None
        for aliases in HEADER_FIELD_ALIASES[:4]
    )


def _records_da_file(file_obj):
    nome_file = file_obj.name.lower()
    tipo_file = _inferisci_tipo_import(nome_file)
    file_obj.seek(0)

    if nome_file.endswith('.csv'):
        text_file = io.TextIOWrapper(file_obj.file, encoding='utf-8-sig')
        records = []
        for row_number, row in enumerate(csv.DictReader(text_file), start=2):
            record = {_normalizza_colonna(key): value for key, value in row.items()}
            record['import_kind'] = tipo_file
            record['source_row'] = row_number
            records.append(record)
        return records

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('per importare file Excel serve installare openpyxl. In alternativa carica un CSV.') from exc

    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    records = []
    for sheet in workbook.worksheets:
        header_row_number, headers = _trova_riga_header(sheet)
        if not headers:
            continue
        tipo_sheet = _inferisci_tipo_import(nome_file, sheet.title) or tipo_file
        righe_vuote_consecutive = 0
        righe_dati_iniziate = False

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row_number + 1, values_only=True),
            start=header_row_number + 1,
        ):
            if _riga_vuota(row):
                if righe_dati_iniziate:
                    righe_vuote_consecutive += 1
                    if righe_vuote_consecutive >= 60:
                        break
                continue

            righe_vuote_consecutive = 0
            record = dict(zip(headers, row))
            record['import_kind'] = tipo_sheet
            record['source_row'] = row_number
            if _record_utilizzabile(record):
                righe_dati_iniziate = True
                records.append(record)
    return records


def _tipo_record_import(row):
    if _valore(row, *IS_COMMESSA_KEYS) is not None:
        return 'commessa' if _bool_import(row, *IS_COMMESSA_KEYS) else 'offerta'
    if row.get('import_kind') in {'commessa', 'offerta'}:
        return row['import_kind']
    if _valore(row, *CODICE_COMMESSA_KEYS) is not None and _valore(row, *CODICE_OFFERTA_KEYS) is None:
        return 'commessa'
    return 'offerta'


def _codice_import(row, tipo_record):
    if tipo_record == 'commessa':
        return _valore(row, *CODICE_GENERICO_KEYS, *CODICE_COMMESSA_KEYS, *CODICE_OFFERTA_KEYS)
    return _valore(row, *CODICE_GENERICO_KEYS, *CODICE_OFFERTA_KEYS, *CODICE_COMMESSA_KEYS)


def _importa_offerte_commesse(file_obj):
    records = _records_da_file(file_obj)

    creati = 0
    aggiornati = 0
    saltati = 0
    codici_processati = set()

    for row in records:
        row = {_normalizza_colonna(key): value for key, value in row.items()}
        tipo_record = _tipo_record_import(row)
        codice = _codice_import(row, tipo_record)
        produttore = _valore(row, *PRODUTTORE_KEYS)
        tipologia = _valore(row, *TIPOLOGIA_KEYS)
        quantita = _int_import(row, *QUANTITA_KEYS)
        codice_normalizzato = str(codice or '').strip()

        if not codice_normalizzato or not produttore or not tipologia or quantita is None:
            saltati += 1
            continue

        if codice_normalizzato in codici_processati:
            continue
        codici_processati.add(codice_normalizzato)

        localita = _valore(row, *LOCALITA_KEYS)
        provincia = _valore(row, *PROVINCIA_KEYS)
        comune = trova_comune(localita, provincia) if localita else None
        is_commessa = tipo_record == 'commessa'
        is_done = _bool_import(row, *IS_DONE_KEYS, default=False)
        if is_commessa and _valore(row, *IS_DONE_KEYS) is None:
            is_done = _bool_import(row, *PROCESSO_COMPLETATO_KEYS, default=False)

        dati = {
            'produttore': str(produttore).strip(),
            'garanzia_fin': _normalizza_garanzia(row),
            'quantita': quantita,
            'tipologia': str(tipologia).strip(),
            'note': str(_valore(row, 'note') or '').strip(),
            'latitudine': _float_import(row, 'latitudine', 'lat'),
            'longitudine': _float_import(row, 'longitudine', 'lon', 'lng'),
            'is_commessa': is_commessa,
            'is_done': is_done,
        }
        if comune:
            dati['paese'] = comune

        _, creato = OffertaCommessa.objects.update_or_create(
            codice=codice_normalizzato,
            defaults=dati,
        )
        creati += int(creato)
        aggiornati += int(not creato)

    return creati, aggiornati, saltati


class AppLoginView(LoginView):
    template_name = 'registration/login.html'
    redirect_authenticated_user = True

    def _is_login_redirect(self, redirect_to):
        return urlparse(redirect_to).path == self.request.path

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        redirect_to = context.get(self.redirect_field_name)
        context['safe_next'] = redirect_to if redirect_to and not self._is_login_redirect(redirect_to) else ''
        return context

    def get_success_url(self):
        redirect_to = self.get_redirect_url()
        if redirect_to and not self._is_login_redirect(redirect_to):
            return redirect_to
        return resolve_url(settings.LOGIN_REDIRECT_URL)


def applica_geolocalizzazione(record, localita=None, provincia=None):
    if record.get('latitudine') and record.get('longitudine'):
        return record

    localita = localita or record.get('localita') or record.get('luogo_ritiro') or record.get('paese__name')
    provincia = provincia or record.get('provincia') or record.get('paese__provincia')
    comune = trova_comune(localita, provincia)
    if not comune:
        return record

    record['latitudine'] = comune.latitudine
    record['longitudine'] = comune.longitudine
    record['paese__latitudine'] = comune.latitudine
    record['paese__longitudine'] = comune.longitudine
    record['paese__name'] = comune.name
    record['paese__provincia'] = comune.provincia
    return record


def anno_da_codice(codice):
    match = re.search(r'(?:^|[_\-\s])(\d{2})\d{3}', str(codice or ''))
    if not match:
        return ''
    anno = 2000 + int(match.group(1))
    return anno if 2000 <= anno <= 2099 else ''


def prepara_record_tabella(record):
    record = applica_geolocalizzazione(record)
    is_commessa = bool(record.get('is_commessa'))
    is_done = bool(record.get('is_done'))
    record['tipo_record'] = 'Commessa' if is_commessa else 'Offerta'
    record['stato_record'] = 'Evasa' if is_done else ('In carico' if is_commessa else 'Offerta')
    record['anno_record'] = anno_da_codice(record.get('codice'))
    return record


class ComuneAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return Comune.objects.none()

        qs = Comune.objects.all()
        if self.q:
            qs = qs.filter(name__icontains=self.q) | qs.filter(cap__istartswith=self.q) | qs.filter(provincia__istartswith=self.q)
        return qs


@login_required
def home(request):
    fornitori = list(Fornitore.objects.values(
        'id',
        'ragione_sociale',
        'indirizzo',
        'trasporto',
        'trattamento',
        'latitudine',
        'longitudine',
        'paese__cap',
        'paese__latitudine',
        'paese__longitudine',
        'paese__name',
        'paese__provincia',
    ))
    commesse = list(Commessa.objects.filter(is_done=False).values(
        'id',
        'codice',
        'produttore',
        'garanzia_fin',
        'quantita',
        'note',
        'tipologia',
        'latitudine',
        'longitudine',
        'paese__latitudine',
        'paese__longitudine',
        'paese__name',
        'paese__provincia',
    ))
    offerte = list(OffertaCommessa.objects.filter(is_commessa=False).values(
        'id',
        'codice',
        'produttore',
        'garanzia_fin',
        'quantita',
        'note',
        'tipologia',
        'latitudine',
        'longitudine',
        'paese__latitudine',
        'paese__longitudine',
        'paese__name',
        'paese__provincia',
    ))
    dati_tabella = list(OffertaCommessa.objects.all().values(
        'id',
        'codice',
        'produttore',
        'garanzia_fin',
        'quantita',
        'note',
        'tipologia',
        'latitudine',
        'longitudine',
        'is_commessa',
        'is_done',
        'paese__cap',
        'paese__latitudine',
        'paese__longitudine',
        'paese__name',
        'paese__provincia',
    ))
    fornitori = [applica_geolocalizzazione(fornitore) for fornitore in fornitori]
    commesse = [applica_geolocalizzazione(commessa) for commessa in commesse]
    offerte = [applica_geolocalizzazione(offerta) for offerta in offerte]
    dati_tabella = [prepara_record_tabella(record) for record in dati_tabella]

    context = {
        'fornitori_list': fornitori,
        'commesse_list': commesse,
        'offerte_list': offerte,
        'dati_tabella_list': dati_tabella,
        'upload_form': ExcelUploadForm(),
        'open_upload_panel': request.GET.get('upload') == '1',
    }
    return render(request, 'HomePage.html', context)


@login_required
def carica_excel(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            files = form.cleaned_data['file']
            creati_totali = 0
            aggiornati_totali = 0
            saltati_totali = 0
            errori = []

            for file_obj in files:
                try:
                    creati, aggiornati, saltati = _importa_offerte_commesse(file_obj)
                    creati_totali += creati
                    aggiornati_totali += aggiornati
                    saltati_totali += saltati
                except Exception as exc:
                    errori.append(f'{file_obj.name}: {exc}')

            if creati_totali or aggiornati_totali or saltati_totali:
                messages.success(
                    request,
                    f'Import completato: {creati_totali} creati, {aggiornati_totali} aggiornati, {saltati_totali} righe saltate.'
                )
            for errore in errori:
                messages.error(request, f'Import non riuscito: {errore}')
            if not (creati_totali or aggiornati_totali or saltati_totali) and not errori:
                messages.error(request, 'Nessuna riga valida trovata nel file caricato.')
            return redirect(f'{resolve_url("home")}?upload=1')

        messages.error(request, 'Seleziona un file Excel o CSV valido.')
        return redirect(f'{resolve_url("home")}?upload=1')

    return redirect(f'{resolve_url("home")}?upload=1')
